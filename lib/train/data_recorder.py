#!/usr/bin/env python
# coding=utf-8
import os
import pandas as pd
import threading
import time
#import random
import h5py
#import numpy as np
#import torch
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
#import copy
import glob  # Import glob for file pattern matching

# --- Configuration ---
_chunk_size = 5                                                                     # Save every 10,000 samples
_delete_chunks_after_merge = True  # Set to False to keep intermediate chunk files
select_sampling = False
# --- Global State (Protected by Lock) ---
_buffer = []
_chunk_files = []
_samples_in_buffer = 0
_total_samples_logged_this_epoch = 0
current_epoch = None
_file_lock = threading.RLock()
sample_per_epoch=0
selected_sampling_epoch=0
mysettings=None
# Define headers based on the original structure
_headers = [
    "Index", "Sample Index", "stats/Loss_total", "stats_IoU", "Seq Name",
    "Template Frame ID", "Template Frame Path", "Search Frame ID", "Seq ID",
    "Seq Path", "Class Name", "Vid ID", "Search Names", "Search Path"
]
# --- Filename Generation ---
def _get_chunk_filename(epoch, start_index, end_index):
    # Save in the root directory where the script is run
    global select_sampling,selected_sampling_epoch
    if select_sampling and selected_sampling_epoch==epoch:
        return f'sample_stats_epoch_{epoch}_all_selected_chunk_sample_{start_index}_{end_index}.xlsx'
    else:
        return f'sample_stats_epoch_{epoch}_all_chunk_sample_{start_index}_{end_index}.xlsx'

def _get_final_filename(epoch, total_samples):
    # Save in the root directory where the script is run
    global select_sampling
    if select_sampling and epoch>=mysettings.selected_sampling_epoch:
        return f'sample_stats_epoch_{epoch}_all_selected_sample_1_{total_samples}.xlsx'
    else:
        return f'sample_stats_epoch_{epoch}_all_sample_1_{total_samples}.xlsx'

def _get_final_filename_unselected(epoch, total_samples):
        return f'sample_stats_epoch_{epoch}_all_sample_1_{total_samples}.xlsx'

# --- Helper Functions ---
def _safe_str_list(value):
    """Safely convert lists or other types to string."""
    if isinstance(value, list):
        # Handle nested lists if necessary, assuming simple list of strings/numbers for now
        return ", ".join(map(str, value))
    elif value is None:
        return ""
    else:
        return str(value)

def _format_excel_file(filename):
    """Applies basic formatting (alignment, column width) to an Excel file."""
    try:
        from openpyxl import load_workbook  # Import locally to avoid dependency if not used
        wb = load_workbook(filename)
        ws = wb.active

        # Center alignment for all cells
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = align

        # Auto-adjust column widths based on content (max length)
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            # Check header length first
            header_cell = ws.cell(row=1, column=col_idx)
            if header_cell.value:
                max_length = len(str(header_cell.value))
            # Check content length (sample a few rows for efficiency if needed)
            for cell in column_cells[1:]:  # Skip header
                try:
                    if cell.value is not None:
                        cell_len = len(str(cell.value))
                        if cell_len > max_length:
                            max_length = cell_len
                except:  # Handle potential errors with cell values
                    pass
            # Add padding
            adjusted_width = max_length + 4
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Set header row height
        ws.row_dimensions[1].height = 25
        wb.save(filename)
    except ImportError:
        print("Warning: openpyxl not found. Cannot apply Excel formatting.",flush=True)
    except Exception as e:
        print(f"Error formatting Excel file {filename}: {e}",flush=True)

# --- Core Logic ---
def _save_chunk(epoch, start_index, end_index, data_to_save):
    """Saves the current buffer to a chunk file."""
    global _chunk_files
    if not data_to_save:
        print("No data in buffer to save as chunk.",flush=True)
        return

    filename = _get_chunk_filename(epoch, start_index, end_index)
    print(f"Saving chunk {start_index}-{end_index} for epoch {epoch} to {filename}...",flush=True)
    try:
        df = pd.DataFrame(data_to_save)
        # Ensure columns are in the correct order
        df = df.reindex(columns=_headers)

        # Save using pandas, specifying the engine
        df.to_excel(filename, index=False, engine='openpyxl')
        # Apply formatting after saving
        _format_excel_file(filename)
        _chunk_files.append(filename)
    except Exception as e:
        print(f"Error saving chunk {filename}: {e}",flush=True)

def _clean_previous_experiments():
    """
    Cleans up all previous experiment files (both chunk and final files).
    This is called at the start of the first epoch.
    """
    global select_sampling
    print("Cleaning up previous experiment files...",flush=True)

    # Define file patterns to search for
    chunk_pattern = "sample_stats_epoch_*_all*_chunk_sample_*.xlsx"
    final_pattern = "sample_stats_epoch_*_all*_sample_*.xlsx"

    # Find all matching files
    existing_files = glob.glob(chunk_pattern) + glob.glob(final_pattern)

    # Print the list of existing files
    if existing_files:
        print("\nFound the following log files:",flush=True)
        for i, f in enumerate(existing_files, 1):
            print(f"  {i}. {os.path.basename(f)}",flush=True)
        print(f"\nTotal files found: {len(existing_files)}",flush=True)

        # Delete all found files if not in select_sampling mode
        if not select_sampling:
            print("\nDeleting Excel files...",flush=True)
            deleted_count = 0
            for f in existing_files:
                try:
                    os.remove(f)
                    print(f"  - Deleted: {os.path.basename(f)}",flush=True)
                    deleted_count += 1
                except OSError as e:
                    print(f"  - Error deleting {os.path.basename(f)}: {e}",flush=True)
            print(f"\nDeletion complete. Total files deleted: {deleted_count}",flush=True)
        else:
            print("\nSelected Sampling is ENABLED. Existing log files will be PRESERVED.",flush=True)
    else:
        print("\nNo existing log files found.",flush=True)
def _merge_chunks(epoch, total_samples):
    """
    Merges all chunk files for the current epoch into a single file.

    Args:
        epoch: Current epoch number
        total_samples: Total number of samples in the epoch
    """
    global _chunk_files

    if not _chunk_files:
        print("No chunk files to merge.",flush=True)
        return

    print(f"Merging {len(_chunk_files)} chunk files for epoch {epoch}...",flush=True)
    all_data_frames = []

    for chunk_file in _chunk_files:
        try:
            df = pd.read_excel(chunk_file, engine='openpyxl')
            all_data_frames.append(df)
        except Exception as e:
            print(f"Error reading chunk file {chunk_file}: {e}. Skipping this chunk.",flush=True)

    if not all_data_frames:
        print("Error: No valid data found in chunk files. Cannot merge.",flush=True)
        return

    # Concatenate all dataframes
    final_df = pd.concat(all_data_frames, ignore_index=True)
    final_df = final_df.reindex(columns=_headers)

    # Save the final merged file
    final_filename = _get_final_filename(epoch, total_samples)
    try:
        final_df.to_excel(final_filename, index=False, engine='openpyxl')
        _format_excel_file(final_filename)
        print(f"Successfully merged chunks into final file: {final_filename}",flush=True)
    except Exception as e:
        print(f"Error saving final merged file: {e}",flush=True)


def _cleanup_chunk_files():
    """Removes all chunk files after they've been merged."""
    global _chunk_files

    if not _chunk_files:
        return

    print("Cleaning up chunk files...",flush=True)
    deleted_count = 0
    for chunk_file in _chunk_files:
        try:
            if os.path.exists(chunk_file):
                os.remove(chunk_file)
                deleted_count += 1
        except Exception as e:
            print(f"Error removing chunk file {chunk_file}: {e}",flush=True)

    print(f"Removed {deleted_count} chunk files.",flush=True)
    _chunk_files = []


def set_sampling(ss):
    global select_sampling
    select_sampling = ss
    print(f"Selected sampling mode set to: {select_sampling}",flush=True)


def set_epoch( settings):
    """
    Sets the current epoch, clearing buffers and state for the new epoch.
    If settings.epoch is 1, also cleans up any previous experiment files.
    """
    global current_epoch, _buffer, _samples_in_buffer, _chunk_files, _total_samples_logged_this_epoch,sample_per_epoch,selected_sampling_epoch,mysettings

    with _file_lock:
        # Clean up previous experiments if this is the first epoch pppp
        if settings.epoch == 1 or current_epoch is None:
            sample_per_epoch = settings.sample_per_epoch
            _clean_previous_experiments()

        print(f"Setting data recorder for epoch {settings.epoch}. Clearing state.",flush=True)
        current_epoch = settings.epoch
        _buffer = []
        _samples_in_buffer = 0
        _chunk_files = []
        _total_samples_logged_this_epoch = 0
        mysettings = settings

def samples_stats_save(sample_index: int, data_info: dict, stats: dict):
    """
    Save sample statistics to the buffer for later logging to Excel.

    Args:
        sample_index: Index of the current sample
        data_info: Dictionary containing sample information
        stats: Dictionary containing sample statistics
    """
    global _buffer, _samples_in_buffer, _total_samples_logged_this_epoch, _chunk_files

    # Determine epoch (should be set by trainer via set_epoch or passed in data_info)
    epoch = current_epoch
    if epoch is None:
        print("Error: Epoch not set in data_recorder. Cannot log data. Call set_epoch() first.",flush=True)
        return

    with _file_lock:
        _total_samples_logged_this_epoch += 1
        current_log_index = _total_samples_logged_this_epoch

        # Create the dictionary matching the headers
        log_entry = {
            "Index": current_log_index,
            "Sample Index": sample_index,
            "stats/Loss_total": stats.get("Loss/total", None),
            "stats_IoU": stats.get("IoU", None),
            "Seq Name": data_info.get("seq_name", ""),
            "Template Frame ID": _safe_str_list(data_info.get("template_ids")),
            "Template Frame Path": _safe_str_list(data_info.get("template_path")),
            "Search Frame ID": _safe_str_list(data_info.get("search_id")),
            "Seq ID": data_info.get("seq_id", ""),
            "Seq Path": data_info.get("seq_path", ""),
            "Class Name": data_info.get("class_name", ""),
            "Vid ID": data_info.get("vid_id", ""),
            "Search Names": _safe_str_list(data_info.get("search_names")),
            "Search Path": _safe_str_list(data_info.get("search_path"))
        }

        _buffer.append(log_entry)
        _samples_in_buffer += 1

        # Save chunk if buffer is full or if this is the last chunk
        if _samples_in_buffer >= _chunk_size or _samples_in_buffer >= mysettings.top_selected_samples:
            start_index = current_log_index - _samples_in_buffer + 1
            end_index = current_log_index
            _save_chunk(epoch, start_index, end_index, _buffer)
            _buffer = []
            _samples_in_buffer = 0

            # Merge and finalize if this is the last sample
        if current_log_index == sample_per_epoch or current_log_index == mysettings.top_selected_samples:
            if _samples_in_buffer > 0:
                start_index = current_log_index - _samples_in_buffer + 1
                end_index = current_log_index
                _save_chunk(epoch, start_index, end_index, _buffer)
            _buffer = []
            _samples_in_buffer = 0
            _merge_chunks(epoch, _total_samples_logged_this_epoch)
            _cleanup_chunk_files()


def save_gradients(model, sample_index, epoch, output_dir='gradients'):
    """
    Save model gradients to an HDF5 file.

    Args:
        model: The PyTorch model
        sample_index: Index of the current sample
        epoch: Current epoch number
        output_dir: Directory to save gradient files
    """
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    # Create HDF5 file for this epoch if it doesn't exist
    h5_file = os.path.join(output_dir, f'gradients_epoch_{epoch}.h5')

    with h5py.File(h5_file, 'a') as f:  # 'a' mode allows appending to existing file
        # Create a group for this sample
        sample_grp = f.create_group(f'sample_{sample_index}')

        # Save gradients for each parameter
        for name, param in model.named_parameters():
            if param.grad is not None:
                # Convert gradient to numpy array and save
                grad_data = param.grad.detach().cpu().numpy()

                # Replace any problematic characters in parameter name that might cause issues with HDF5
                safe_name = name.replace('.', '_')

                # Save gradient data
                sample_grp.create_dataset(safe_name, data=grad_data, compression='gzip')

                # Also save some metadata
                sample_grp.attrs[f'{safe_name}_shape'] = str(param.grad.shape)
                sample_grp.attrs[f'{safe_name}_dtype'] = str(param.grad.dtype)

        # Save timestamp
        sample_grp.attrs['timestamp'] = time.strftime('%Y-%m-%d %H:%M:%S')
        sample_grp.attrs['sample_index'] = sample_index
        sample_grp.attrs['epoch'] = epoch